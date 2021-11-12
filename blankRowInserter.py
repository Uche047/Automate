#! python3
# blankRowInserter.py The program inserts blank rows into .xlsx files with the row to start after 
# provided in the command line, number of rows to add and the .xlsx file in that order 
# Making neccessary imports
import openpyxl, os, sys
# Changing directory
os.chdir("C:\\Users\\patri\\PycharmProjects\\Developper")
# Loading the .xlsx file in a workbook object
wb = openpyxl.load_workbook(sys.argv[3])
# Grabing the active sheet
sheet = wb.active
# Converting the row number to start after and rows to insert from string to integer format 
N = int(sys.argv[1])
M = int(sys.argv[2])
# Adding the number of rows to be inserted to maximum number of rows
# To accommodate for the extra rows
n = int(sheet.max_row) + M
# Subtracting the number of rows from n for the for loop
l = n - M -1
# Looping through provided row number is less than row to insert blank rows after

while l > N:
    for row in range(1, sheet.max_row + N + 2):
        for column in range(1, sheet.max_row + 1):
            sheet.cell(row = n, column = column).value = sheet.cell(row = l, column = column).value
    # Decreasing by 1 to properly assign values to lower row numbers
    n -= 1
    l -= 1
    
# Loop through to insert blank rows
for row in range(N, N+M):
    for column in range(1, sheet.max_row + 1):
        sheet.cell(row = row, column = column).value = " "

# Saving your file with a new prefix updated        
wb.save('updated_' + sys.argv[3])