#! python3
# multiplicationTable.py takes an integer from the command line
# And creates a multiplication table saved in .xlsx format
# Making the necessary imports
import openpyxl, os,sys
# Changing Directory
os.chdir("C:\\Users\\patri\\PycharmProjects\\Developper")
# Importing Font
from openpyxl.styles import Font
# Creating a new workbook
wb = openpyxl.Workbook()
# Working with the active sheet
sheet = wb.active
myBold = Font(bold=True)
# Looping through and setting values bold
for i in range(1, int(sys.argv[1]) + 1):
    sheet.cell(row=1, column=i+1).font = myBold
    sheet.cell(row=1, column=i+1).value = i
    sheet.cell(row=i+1, column=1).font = myBold
    sheet.cell(row=i+1, column=1).value = i
# Looping through and doing multiplication calculations
for i in range(2, int(sys.argv[1]) + 2 ):
    for j in range(2, int(sys.argv[1]) + 2):
        sheet.cell(row=i, column=j).value = sheet.cell(row=i, column=1).value * sheet.cell(row=1, column=j).value

# Saving workbook
wb.save('Multiplication.xlsx')
    